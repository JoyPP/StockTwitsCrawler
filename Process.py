import os
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

def compare_datasets(old_dir = 'original_dataset/', new_dir = 'single_symbol_dataset/'):
    if os.path.exists(old_dir):
        files = [f for f in os.listdir(old_dir) if f.endswith('.xlsx')]
    num_list = "File\tBefore\tAfter\n"
    before = 0
    after = 0
    for f in files:
        print 'file ', f
        wb = load_workbook(filename=old_dir+f)
        ws = wb.get_active_sheet()
        row1 = ws.max_row - 1
        before += row1

        wb = load_workbook(filename=new_dir + f)
        ws = wb.get_active_sheet()
        row2 = ws.max_row - 1
        after += row2

        num_list += f + '\t' + str(row1) + '\t' + str(row2) + '\n'

    num_list += 'Total\t' + str(before) + '\t' + str(after) + '\n'
    print 'before = ', before, 'after = ', after
    with open('statistic.txt','w') as f:
        f.write(num_list)

def copy_row(ws, tgt_ws, row):
    col = ws.max_column
    if row == 1:
        tgt_row = 1
    else:
        tgt_row = tgt_ws.max_row+1
    col_dict = {1:'A', 2:'B', 3:'C',4:'D', 5:'E', 6:'F', 7: 'G', 8:'H', 9:'I', 10:'J',11:'K',12:'L',13:'M',14:'N',15:'O',16:'P',17:'Q',
                18:'R',19:'S',20:'T',21:'U',22:'V',23:'W',24:'X',25:'Y',26:'Z'}
    print 'tgt row = ', tgt_row
    for j in range(1, col+1):
        tgt_ws.cell(row=tgt_row, column=j, value=ws[col_dict[j]+str(row)].value)


def keywords_selection(old_dir = 'single_symbol_dataset/', new_dir = 'sample/'):
    if os.path.exists(old_dir):
        files = [f for f in os.listdir(old_dir) if f.endswith('.xlsx')]
    keywords = ['would','will','buy','sell','keep','hold','up','down','predict','expect']
    for f in files:
        wb = load_workbook(filename=old_dir + f)
        ws = wb.get_active_sheet()
        row = 21#ws.max_row

        tgt_wb = Workbook()
        tgt_ws = tgt_wb.active
        tgt_ws.title = f[:-5]

        copy_row(ws, tgt_ws, 1)

        for i in range(2, row+1):
            msg = ws['B'+str(i)].value.encode('utf-8')
            flag = False
            for key in keywords:
                if msg.find(key) >= 0:
                    flag = True
            if flag:
                copy_row(ws, tgt_ws, i)

        tgt_wb.save(new_dir+f)


keywords_selection()