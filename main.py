import json
import urllib2
from bs4 import BeautifulSoup
import time
import cPickle
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
import os
import re
import argparse

def save_message(symbol, messages):
    if not os.path.exists('messages/'):
        os.mkdir('messages/')
    filename = 'messages/' + symbol + '.xlsx'
    print 'save content into', filename
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
    if os.path.exists(filename):
        wb = load_workbook(filename=filename)
        ws = wb[symbol]
        row = ws.max_row
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = symbol
        ws.cell(row=1, column=1, value='MessageID')
        ws.cell(row=1, column=2, value='Message')
        ws.cell(row=1, column=3, value='Created_time')
        ws.cell(row=1, column=4, value='Context')
        ws.cell(row=1, column=5, value='Symbol')
        ws.cell(row=1, column=6, value='Sentiment')
        ws.cell(row=1, column=7, value='Msg_Like_Count')
        ws.cell(row=1, column=8, value='Mes_Like_IDs')
        ws.cell(row=1, column=9, value='Username')
        ws.cell(row=1, column=10, value='Name')
        ws.cell(row=1, column=11, value='UserID')
        ws.cell(row=1, column=12, value='Official')
        ws.cell(row=1, column=13, value='Classification')
        ws.cell(row=1, column=14, value='User_Like_Count')
        ws.cell(row=1, column=15, value='Ideas')
        ws.cell(row=1, column=16, value='Followers')
        ws.cell(row=1, column=17, value='Following')
        ws.cell(row=1, column=18, value='Join_Date')
        ws.cell(row=1, column=19, value='Watchlist_Stocks_Count')
        ws.cell(row=1, column=20, value='Identity')
        row = 1
    for i, msg in zip(range(row + 1, row + 1 + len(messages)), messages):
        ws.cell(row=i, column=1, value=msg['id'])
        try:
            ws.cell(row=i, column=2, value=msg['body'])
        except:
            ws.cell(row=i, column=2, value=ILLEGAL_CHARACTERS_RE.sub(r'', msg["body"]))
        ws.cell(row=i, column=3, value=msg['created_at'])
        try:
            ws.cell(row=i, column=4, value=msg['context'])
        except:
            ws.cell(row=i, column=4, value=ILLEGAL_CHARACTERS_RE.sub(r'',msg['context']))
        ws.cell(row=i, column=5, value=msg['symbols'])
        ws.cell(row=i, column=6, value=msg['sentiment'])
        ws.cell(row=i, column=7, value=msg['likes_count'])
        ws.cell(row=i, column=8, value=msg['likes_id'])
        ws.cell(row=i, column=9, value=msg['user']['username'])
        ws.cell(row=i, column=10, value=msg['user']['name'])
        ws.cell(row=i, column=11, value=msg['user']['id'])
        ws.cell(row=i, column=12, value=msg['user']['official'])
        ws.cell(row=i, column=13, value=msg['user']['classification'])
        ws.cell(row=i, column=14, value=msg['user']['like_count'])
        ws.cell(row=i, column=15, value=msg['user']['ideas'])
        ws.cell(row=i, column=16, value=msg['user']['followers'])
        ws.cell(row=i, column=17, value=msg['user']['following'])
        ws.cell(row=i, column=18, value=msg['user']['join_date'])
        ws.cell(row=i, column=19, value=msg['user']['watchlist_stocks_count'])
        ws.cell(row=i, column=20, value=msg['user']['identity'])
    wb.save(filename)

def parse_message(symbol, messages):
    print 'parse %d messages' % len(messages)

    for i, msg in enumerate(messages):
        #print msg

        keys = msg.keys()
        #body = msg["body"]  # message information

        context = ""
        if "links" in keys:  # links mentioned, can be used as context
            links = msg["links"]
            for link in links:
                if "title" in link.keys() and link["title"] is not None:
                    context += link["title"]
                elif "description" in link.keys() and link["description"] is not None:
                    context += link["description"]
            messages[i].pop('links')
        messages[i]["context"] = context

        sentiment = msg["entities"]["sentiment"]  # author's sentiment
        if sentiment is not None:
            sentiment = sentiment["basic"]
        else:
            sentiment = ""
        messages[i]["sentiment"] = sentiment
        messages[i].pop('entities')

        #created_time = msg["created_at"]  # msg created time
        #mentioned_users = msg["mentioned_users"]  # mentioned users
        #mentioned_symbols = msg["symbols"]  # mentioned symbols
        #msg_id = msg["id"]  # msg id
        mentioned_symbols = list()
        for s in msg["symbols"]:
            mentioned_symbols.append(s['symbol'])
        msg["symbols"] = ";".join(mentioned_symbols)

        if "likes" in keys:
            likes = msg["likes"]
            likes_id = ";".join(map(str, likes["user_ids"]))
            likes_count = likes["total"]
        else:
            likes_id = ""
            likes_count = 0
        messages[i]['likes_id'] = likes_id
        messages[i]['likes_count'] = likes_count
        messages[i].pop('likes', None)

        # user information
        messages[i]['user']['classification'] = ';'.join(msg['user']['classification'])
        '''
        user = msg["user"]
        username = user["username"]
        name = user["name"]
        classification = user["classification"]  # list: [], ["suggested"], ["suggested", "official"]
        official = user["official"]
        like_count = user["like_count"]  # likes received in total
        ideas = user["ideas"]
        following = user["following"]
        join_date = user["join_date"]
        follower = user["followers"]
        watchlist_stocks_count = user["watchlist_stocks_count"]
        user_id = user["id"]
        identity = user["identity"]  # User, Official
        '''
    save_message(symbol, messages)
    return msg["id"]   # return the last id, for next crawling


def get_messages(symbol, max_epoch, start_id = None):
    url = 'https://api.stocktwits.com/api/2/streams/symbol/' + symbol + '.json?filter=top&max'
    if id is not None:
        url += "=" + str(start_id)
    hdr = {'User-Agent': 'Mozilla/5.0'}
    print 'Request url =', url
    req = urllib2.Request(url, headers=hdr)
    html = urllib2.urlopen(req).read()
    start = time.time()
    for i in range(max_epoch):
        try:
            js = json.loads(str(html))
        except:
            js = None
        if js is not None:
            last_id = parse_message(symbol, js["messages"])
            url = 'https://api.stocktwits.com/api/2/streams/symbol/' + symbol + '.json?filter=top&max=' + str(last_id)
            print 'Request url =', url
            while True:
                try:
                    req = urllib2.Request(url, headers=hdr)
                    html = urllib2.urlopen(req).read()
                    break
                except:
                    time.sleep(3)
    print 'time spend = ', time.time() - start


parser = argparse.ArgumentParser(description='StockTwits Crawler')
parser.add_argument('--symbol', default='', type=str, help='stock symbol')
parser.add_argument('--max_epoch', default=3000, type=int, help='#pages to parse')
args = parser.parse_args()
symbol = args.symbol
start_id = None
filename = 'messages/' + symbol + '.xlsx'
if os.path.exists(filename):
    wb = load_workbook(filename=filename)
    ws = wb[symbol]
    row = ws.max_row
    start_id = int(ws['A'+str(row)].value)
print symbol, ' start id =', start_id
get_messages(symbol, args.max_epoch, start_id=start_id)
